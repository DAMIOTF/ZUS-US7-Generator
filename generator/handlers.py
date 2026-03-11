# -*- coding: utf-8 -*-
"""
Logika obsługi zdarzeń i generowania PDF.

Klasa ``HandlersMixin`` jest wmiksowana do ``App`` i zawiera
metody obsługi plików, logowania, zbierania opcji oraz generowania PDF.
"""
import datetime
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import Dict, Optional

import openpyxl

from .constants import (
    C_GREEN, C_RED, C_SUBTEXT, C_YELLOW,
)
from .excel_loader import load_sheet_data_mapped
from .pdf_engine import fill_pdf


class HandlersMixin:
    """Mixin — logika obsługi zdarzeń i generowania PDF."""

    # =========================================================================
    # Log helpers
    # =========================================================================

    def _log_write(self, msg: str, tag: str = "") -> None:
        self._log.configure(state="normal")
        self._log.insert("end", msg + "\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")

    def _clear_log(self) -> None:
        self._log.configure(state="normal")
        self._log.delete("1.0", "end")
        self._log.configure(state="disabled")

    def _set_status(self, msg: str, color: str = C_SUBTEXT) -> None:
        self._status_lbl.configure(text=msg, fg=color)

    def _update_stats(
        self,
        total=None,  # type: Optional[int]
        ok=None,     # type: Optional[int]
        err=None,    # type: Optional[int]
    ) -> None:
        if total is not None:
            self._stat_total.configure(text=f"Wierszy: {total}")
        if ok is not None:
            self._stat_ok.configure(text=f"OK: {ok}")
        if err is not None:
            self._stat_err.configure(text=f"Błędy: {err}")

    # =========================================================================
    # Obsługa pliku / arkusza
    # =========================================================================

    def _browse_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="Wybierz plik Excel",
            filetypes=[("Pliki Excel", "*.xlsx *.xls"), ("Wszystkie", "*.*")])
        if not path:
            return
        try:
            wb     = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as exc:
            messagebox.showerror("Błąd odczytu",
                                 f"Nie można otworzyć pliku:\n{exc}")
            return
        self._wb_path.set(path)
        self._sheet_cb["values"] = sheets
        if sheets:
            self._sheet_var.set(sheets[0])
            self.after(50, self._on_sheet_selected)
        self._log_write(
            f"◈ Wczytano: {os.path.basename(path)}  ({len(sheets)} ark.)", "inf")

    def _on_sheet_selected(self, _=None) -> None:
        sheet = self._sheet_var.get()
        if not sheet or not self._wb_path.get():
            return
        mapping = self._collect_cell_mapping()
        if not mapping:
            self._record_count_lbl.configure(
                text="  ⚠  Podaj komórki mapowania", fg=C_YELLOW)
            self._update_stats(total=0)
            return
        try:
            wb     = openpyxl.load_workbook(
                self._wb_path.get(), read_only=False, data_only=True)
            people = load_sheet_data_mapped(wb, sheet, mapping)
            wb.close()
            n = len(people)
            self._record_count_lbl.configure(
                text=f"  ✔  {n} rekordów",
                fg=C_GREEN if n else C_YELLOW)
            self._update_stats(total=n)
        except Exception:
            self._record_count_lbl.configure(
                text="  ✘  błąd odczytu", fg=C_RED)

    def _browse_output(self) -> None:
        path = filedialog.askdirectory(title="Wybierz folder wynikowy")
        if path:
            self._output_var.set(path)

    def _reset_uzas(self) -> None:
        self._uzas_text.delete("1.0", "end")
        self._uzas_text.insert("1.0", self._uzasadnienie_default)

    # =========================================================================
    # Zbieranie mapowania komórek
    # =========================================================================

    def _collect_cell_mapping(self) -> Dict[str, str]:
        """Zwraca słownik {nazwa_pola: ref_komórki} tylko dla niepustych pól."""
        mapping = {}  # type: Dict[str, str]
        for key, var in self._cell_vars.items():
            val = var.get().strip()
            if val:
                mapping[key] = val
        return mapping

    # =========================================================================
    # Zbieranie opcji PDF z GUI
    # =========================================================================

    def _collect_options(self) -> dict:
        return {
            "radio_type":        self._cv["radio_type"].get(),
            "cb_ubezpieczenia":  self._cv["cb_ubezpieczenia"].get(),
            "cb_przerwy":        self._cv["cb_przerwy"].get(),
            "cb_podstawy":       self._cv["cb_podstawy"].get(),
            "cb_warunki":        self._cv["cb_warunki"].get(),
            "cb_ofe_czlonek":    self._cv["cb_ofe_czlonek"].get(),
            "cb_ofe_skladki":    self._cv["cb_ofe_skladki"].get(),
            "delivery_placowka": self._cv["delivery_placowka"].get(),
            "delivery_poczta":   self._cv["delivery_poczta"].get(),
            "delivery_pue":      self._cv["delivery_pue"].get(),
            "uzasadnienie":      self._uzas_text.get("1.0", "end-1c").strip(),
        }

    # =========================================================================
    # Generowanie PDF
    # =========================================================================

    def _on_generate(self) -> None:
        excel_path = self._wb_path.get().strip()
        sheet_name = self._sheet_var.get().strip()
        date_str   = self._date_var.get().strip()
        output_dir = self._output_var.get().strip()

        if not excel_path:
            messagebox.showwarning("Brak pliku", "Wybierz plik Excel z danymi.")
            return
        if not sheet_name:
            messagebox.showwarning("Brak arkusza", "Wybierz arkusz danych.")
            return
        if not date_str:
            messagebox.showwarning("Brak daty", "Podaj datę wniosku.")
            return

        cell_mapping = self._collect_cell_mapping()
        if not cell_mapping:
            messagebox.showwarning(
                "Brak mapowania",
                "Podaj przynajmniej jedną komórkę startową\n"
                "w sekcji 'Mapowanie komórek Excel'.")
            return

        options = self._collect_options()

        self._btn_generate.set_state("disabled")
        self._btn_generate.configure_text("⏳  Generowanie…")
        self._dot.start()
        self._progress_bar.set_progress(0)
        self._update_stats(ok=0, err=0)
        self._set_status("⏳  Generowanie w toku…", C_YELLOW)

        sep = "─" * 54
        self._log_write(sep, "muted")
        self._log_write(
            f"▶ Start  {datetime.datetime.now().strftime('%H:%M:%S')}  "
            f"│  Arkusz: {sheet_name}  │  Data: {date_str}", "head")

        threading.Thread(
            target=self._generate_worker,
            args=(excel_path, sheet_name, date_str, output_dir,
                  options, cell_mapping),
            daemon=True,
        ).start()

    def _generate_worker(
        self,
        excel_path: str,
        sheet_name: str,
        date_str: str,
        output_dir: str,
        options: dict,
        cell_mapping,  # type: Dict[str, str]
    ) -> None:
        ok_count = err_count = 0
        try:
            wb     = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            people = load_sheet_data_mapped(wb, sheet_name, cell_mapping)
            wb.close()

            if not people:
                self.after(0, lambda: self._log_write(
                    "  Brak rekordów w arkuszu.", "warn"))
                self.after(0, self._done_reset)
                return

            n = len(people)
            self.after(0, lambda: self._log_write(
                f"  Znaleziono {n} rekordów do przetworzenia.", "inf"))
            self.after(0, lambda: self._update_stats(total=n))

            os.makedirs(output_dir, exist_ok=True)

            for idx, person in enumerate(people):
                try:
                    first = person["imie"].split()[0] if person["imie"] else "brak"
                    fname = f"ZUS {first} {person['nazwisko']}.pdf"
                    out   = os.path.join(output_dir, fname)
                    pdf   = fill_pdf(person, date_str, options)
                    with open(out, "wb") as fh:
                        fh.write(pdf)
                    ok_count += 1
                    msg = f"  ✔  {fname}"
                    self.after(0, lambda m=msg: self._log_write(m, "ok"))
                except Exception as exc:
                    err_count += 1
                    name = f"{person.get('imie','')} {person.get('nazwisko','')}"
                    msg  = f"  ✘  {name}: {exc}"
                    self.after(0, lambda m=msg: self._log_write(m, "err"))

                pct  = (idx + 1) / n
                ok_c = ok_count; err_c = err_count
                self.after(0, lambda p=pct, o=ok_c, e=err_c: (
                    self._progress_bar.set_progress(p),
                    self._update_stats(ok=o, err=e),
                ))

            ok_f  = ok_count
            err_f = err_count
            total = n

            def _done() -> None:
                self._progress_bar.animate_to(1.0)
                self._dot.stop()
                self._log_write(
                    f"✔ Zakończono — {ok_f}/{total} plików wygenerowanych"
                    + (f", błędy: {err_f}" if err_f else ""), "head")
                self._log_write(f"  Folder: {output_dir}", "inf")
                self._set_status(
                    f"✔ Gotowe! {ok_f}/{total} plików",
                    C_GREEN if not err_f else C_YELLOW)
                self._btn_generate.set_state(
                    "normal", command=self._btn_generate._orig_command)
                self._btn_generate.configure_text("⚡  Generuj PDF")
                if ok_f:
                    try:
                        os.startfile(output_dir)
                    except Exception:
                        pass

            self.after(0, _done)

        except Exception as exc:
            def _err(e=exc) -> None:
                self._log_write(f"  ✘ Błąd krytyczny: {e}", "err")
                self._set_status("Błąd — sprawdź dziennik", C_RED)
                self._dot.stop()
                self._btn_generate.set_state(
                    "normal", command=self._btn_generate._orig_command)
                self._btn_generate.configure_text("⚡  Generuj PDF")
            self.after(0, _err)

    def _done_reset(self) -> None:
        self._dot.stop()
        self._set_status("Gotowy do generowania", C_SUBTEXT)
        self._btn_generate.set_state(
            "normal", command=self._btn_generate._orig_command)
        self._btn_generate.configure_text("⚡  Generuj PDF")
